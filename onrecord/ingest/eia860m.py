"""EIA-860M generator-inventory ingest — T-063 (module layer).

The monthly EIA-860M workbook is the keyless federal record of what
generation is actually RUNNING and what is actually FILED to come online
— the grid-facts counterpart to hearing-room promises. This module
parses its "Operating" and "Planned" sheets into ONE normalized row
shape:

    {plant, entity, county, state, mw, technology, sector, status,
     planned_online}

(all str except mw float; status is "operating"/"planned" derived from
the SHEET, never from the sheet's own Status column; planned_online is
"" on Operating rows and "YYYY-MM" on Planned rows only when BOTH
Planned Operation Month and Year are present — blanks are literal " "
cells in the real file).

Workbook layout (verified against june_generator2026.xlsx, fetched
2026-08-14): title row, blank row, header on ROW 3 (located by NAME here,
not position), data rows, then a trailing blank row and a "NOTES:"
footer row — rows without a Plant Name are skipped. Structural surprises
(unknown sheet, missing header, bytes that are not an xlsx) raise a LOUD
ValueError: a silent [] on a schema change would quietly zero the
dataset. The operational fetcher degrades to [] instead — same split as
T-066.

County+state joins to registry jurisdictions REUSE iso_queues'
join_jurisdictions/load_mapping and the SAME data/iso_jurisdictions.json
table (same counties, same exact-match honesty: misses returned, never
fuzzy-matched). Rows are structured data, not corpus docs — they never
enter the search index (T-059 honesty pin).

Fetch reality pin (2026-08-15): the index page lists the current file
under /xls/ and history under /archive/xls/, PLUS dead future-month
placeholder links that 301 to an HTML landing page — so the fetcher
walks links newest-first and skips any download that is not actually a
zip (xlsx) payload.
"""

from __future__ import annotations

import io
import logging
import re
from pathlib import Path

import httpx
import openpyxl

from onrecord.ingest.iso_queues import (
    _get_with_retries,
    _snapshot,
    join_jurisdictions,
    load_mapping,
)

__all__ = [
    "parse_860m",
    "latest_links",
    "fetch_860m",
    "join_jurisdictions",
    "load_mapping",
]

logger = logging.getLogger(__name__)

EIA_BASE = "https://www.eia.gov"
INDEX_URL = f"{EIA_BASE}/electricity/data/eia860m/"
DEFAULT_OUT_DIR = Path("artifacts/eia")

_MONTHS = (
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
)
_LINK_RE = re.compile(
    r'href="([^"]*/(' + "|".join(_MONTHS) + r')_generator(\d{4})\.xlsx)"',
    re.IGNORECASE,
)

# status is sheet-derived; only these two sheets are normalized.
_SHEET_STATUS = {"Operating": "operating", "Planned": "planned"}

_HEADER_SCAN_ROWS = 10
_REQUIRED_COLUMNS = (
    "Entity Name", "Plant Name", "Plant State", "County",
    "Nameplate Capacity (MW)", "Technology", "Sector",
)
_PLANNED_COLUMNS = ("Planned Operation Month", "Planned Operation Year")

# The live page (2026-08-15) lists SIX dead future-month placeholders ahead
# of the real current file (july..december 2026, all 301 -> HTML), so the
# newest-first walk needs enough tries to reach the first real workbook.
_MAX_LINK_TRIES = 8


def _s(value: object) -> str:
    """Cell -> stripped str; None -> "" (blank cells are None or " ")."""
    return "" if value is None else str(value).strip()


def _mw(value: object) -> float:
    """MW cell -> float; blank or unreadable -> 0.0 (a filed generator
    with an unreadable MW still counts as a row)."""
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def _planned_online(month: object, year: object) -> str:
    """Planned Operation Month/Year cells -> "YYYY-MM" only when BOTH are
    present (real blanks are literal " " cells); otherwise ""."""
    month_s, year_s = _s(month), _s(year)
    if not (month_s.isdigit() and year_s.isdigit()):
        return ""
    return f"{int(year_s):04d}-{int(month_s):02d}"


def parse_860m(xlsx_bytes: bytes, sheet: str) -> list[dict]:
    """EIA-860M workbook bytes + sheet name -> normalized rows. Sheet must
    be exactly "Operating" or "Planned"; structural surprises are LOUD
    ValueErrors; footer/blank rows (no Plant Name) are skipped."""
    status = _SHEET_STATUS.get(sheet)
    if status is None:
        raise ValueError(f"eia860m: unsupported sheet {sheet!r} (Operating|Planned)")
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ValueError(f"eia860m: not an xlsx workbook ({type(exc).__name__})") from exc
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"eia860m: workbook has no {sheet!r} sheet")
        worksheet = workbook[sheet]

        required = _REQUIRED_COLUMNS + (_PLANNED_COLUMNS if status == "planned" else ())
        rows_iter = worksheet.iter_rows(values_only=True)
        index: dict[str, int] | None = None
        for _ in range(_HEADER_SCAN_ROWS):
            row = next(rows_iter, None)
            if row is None:
                break
            cells = {_s(cell): i for i, cell in enumerate(row) if _s(cell)}
            if all(name in cells for name in required):
                index = {name: cells[name] for name in required}
                break
        if index is None:
            raise ValueError(f"eia860m: header row not found on sheet {sheet!r}")

        def cell(row: tuple, name: str) -> object:
            i = index[name]
            return row[i] if i < len(row) else None

        rows: list[dict] = []
        for raw in rows_iter:
            plant = _s(cell(raw, "Plant Name"))
            if not plant:
                continue  # trailing blank row / NOTES footer / stray rows
            rows.append(
                {
                    "plant": plant,
                    "entity": _s(cell(raw, "Entity Name")),
                    "county": _s(cell(raw, "County")),
                    "state": _s(cell(raw, "Plant State")),
                    "mw": _mw(cell(raw, "Nameplate Capacity (MW)")),
                    "technology": _s(cell(raw, "Technology")),
                    "sector": _s(cell(raw, "Sector")),
                    "status": status,
                    "planned_online": (
                        _planned_online(
                            cell(raw, "Planned Operation Month"),
                            cell(raw, "Planned Operation Year"),
                        )
                        if status == "planned"
                        else ""
                    ),
                }
            )
        return rows
    finally:
        workbook.close()


def latest_links(html: str) -> list[str]:
    """Index-page HTML -> absolute <month>_generator<year>.xlsx URLs,
    newest first by (year, month), deduplicated. Dead future-month
    placeholder links stay in place — the FETCHER verifies payloads."""
    seen: set[str] = set()
    found: list[tuple[int, int, str]] = []
    for href, month, year in _LINK_RE.findall(html):
        url = href if href.startswith("http") else f"{EIA_BASE}{href}"
        if url in seen:
            continue
        seen.add(url)
        found.append((int(year), _MONTHS.index(month.lower()) + 1, url))
    found.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return [url for _, _, url in found]


# --------------------------------------------------------------------------
# Operational fetcher (live network; not unit-tested)
# --------------------------------------------------------------------------


def fetch_860m(out_dir: str | Path = DEFAULT_OUT_DIR) -> list[dict]:
    """Scrape the EIA-860M index for the newest real workbook, snapshot it
    under artifacts/eia/, and return Operating+Planned normalized rows
    ([] on any failure — the fetcher degrades, only the parser is loud).
    Non-zip payloads (the dead future-month placeholder links redirect to
    an HTML landing page) are skipped in favor of the next-newest link."""
    with httpx.Client() as client:
        page = _get_with_retries(client, INDEX_URL)
        if page is None:
            return []
        for url in latest_links(page.text)[:_MAX_LINK_TRIES]:
            response = _get_with_retries(client, url)
            if response is None:
                continue
            if not response.content.startswith(b"PK"):
                logger.debug("eia860m: %s is not an xlsx payload, skipping", url)
                continue
            _snapshot(Path(out_dir), "eia860m.xlsx", response.content, url)
            try:
                return parse_860m(response.content, "Operating") + parse_860m(
                    response.content, "Planned"
                )
            except ValueError:
                logger.warning("eia860m: downloaded workbook failed to parse (%s)", url)
                return []
    return []
