"""ISO interconnection-queue ingest — T-059 (module layer).

Grid operators' interconnection queues are structured, public, keyless
records of what was actually FILED — the counterpart to what was PROMISED
in hearing rooms. This module parses the four verified keyless feeds into
ONE normalized row shape:

    {iso, queue_id, county, state, mw, fuel, status, queue_date, withdrawn}

(all str except mw float and withdrawn bool), then joins county+state to
our registry jurisdictions via an EXPLICIT exact-match table
(data/iso_jurisdictions.json) — unmapped rows are returned as misses,
never fuzzy-matched or dropped. Queue rows are structured data, not
corpus docs: they never enter the search index (T-059 honesty pin).

Sources (verified 2026-08-14, xlsx feeds re-verified 2026-08-15):
- MISO:  https://www.misoenergy.org/api/giqueue/getprojects (JSON list)
- SPP:   https://opsportal.spp.org/Studies/GenerateActiveCSV (CSV with a
  "Last Updated On" preamble line above the header; active requests only,
  so "Date Withdrawn" is empty in practice but still honored)
- ERCOT: the GIS Report workbook via the MIS document listing
  (reportTypeId=15933 — which mixes Co-located Battery reports into the
  same listing, so the newest GIS_Report entry is picked by FriendlyName,
  then downloaded through mirDownload). Sheet "Project Details - Large
  Gen", header row 31. ERCOT rows ALONE carry three additive keys —
  air_permit, ghg_permit, water_availability — verbatim from the sheet.
- CAISO: http://www.caiso.com/PublishedDocuments/PublicQueueReport.xlsx,
  sheet "Grid GenerationQueue", header row 4. NOT CA-only: rows span
  CA/NV/AZ/MX and counties come UPPERCASE.

T-059 AMENDMENT (2026-08-15): openpyxl became a project dependency, so
the original CAISO/ERCOT xlsx scope trim is lifted. PJM still needs keys
or click-through and stays out per the keyless posture. The workbooks are
zipped XML read by openpyxl's stdlib expat parser (entity expansion is
capped there), so the DTD-reject posture (T-038/T-052) needs no extra
code here. Pure parsers pinned by tests/unit/ingest/test_iso_queues.py.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import time
from datetime import UTC, date, datetime
from pathlib import Path

import httpx
import openpyxl

logger = logging.getLogger(__name__)

MISO_URL = "https://www.misoenergy.org/api/giqueue/getprojects"
SPP_URL = "https://opsportal.spp.org/Studies/GenerateActiveCSV"
ERCOT_LIST_URL = "https://www.ercot.com/misapp/servlets/IceDocListJsonWS?reportTypeId=15933"
ERCOT_DOWNLOAD_URL = "https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId={doc_id}"
CAISO_URL = "http://www.caiso.com/PublishedDocuments/PublicQueueReport.xlsx"
HEADERS = {"User-Agent": "OnRecord research alexander.miller@challenger.gauntletai.com"}
DEFAULT_OUT_DIR = Path("artifacts/iso")
MAPPING_PATH = Path("data/iso_jurisdictions.json")

_SPP_KEY_COLUMN = "Generation Interconnection Number"
_ERCOT_SHEET = "Project Details - Large Gen"
_CAISO_SHEET = "Grid GenerationQueue"


def _iso_date(value: str) -> str:
    """Feed date -> YYYY-MM-DD. MISO sends ISO timestamps (keep the date
    part); SPP sends M/D/YYYY; empty stays empty; anything else verbatim."""
    value = (value or "").strip()
    if not value:
        return ""
    if "T" in value:
        return value.split("T", 1)[0]
    parts = value.split("/")
    if len(parts) == 3 and all(p.isdigit() for p in parts) and len(parts[2]) == 4:
        month, day, year = parts
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return value


def _mw(value: object) -> float:
    """MW cell -> float; blank or unparseable -> 0.0 (never a crash — a
    filed project with an unreadable MW still counts as a row)."""
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def parse_miso(json_text: str) -> list[dict]:
    """MISO giqueue getprojects JSON -> normalized rows. Malformed JSON or
    a non-list payload -> []; entries without a projectNumber skipped;
    county/state kept verbatim (multi-county strings are NOT split);
    withdrawn = withdrawnDate set OR applicationStatus == "Withdrawn"."""
    try:
        data = json.loads(json_text)
    except (TypeError, ValueError):
        logger.debug("miso queue: unparseable JSON")
        return []
    if not isinstance(data, list):
        logger.debug("miso queue: expected a JSON list, got %s", type(data).__name__)
        return []

    rows: list[dict] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        queue_id = str(item.get("projectNumber") or "").strip()
        if not queue_id:
            continue
        status = str(item.get("applicationStatus") or "").strip()
        withdrawn_date = str(item.get("withdrawnDate") or "").strip()
        rows.append(
            {
                "iso": "MISO",
                "queue_id": queue_id,
                "county": str(item.get("county") or "").strip(),
                "state": str(item.get("state") or "").strip(),
                "mw": _mw(item.get("summerNetMW")),
                "fuel": str(item.get("fuelType") or "").strip(),
                "status": status,
                "queue_date": _iso_date(str(item.get("queueDate") or "")),
                "withdrawn": bool(withdrawn_date) or status == "Withdrawn",
            }
        )
    return rows


def parse_spp(csv_text: str) -> list[dict]:
    """SPP GenerateActiveCSV text -> normalized rows. The real file has a
    "Last Updated On" preamble above the header and a leading space in the
    " Nearest Town or County" header cell — the header row is located by
    its key column and every header stripped. No header -> []; rows
    without a queue number skipped; fuel = "Generation Type", falling back
    to "Fuel Type" when empty (652 of 1023 live rows have no Fuel Type)."""
    lines = csv_text.splitlines()
    header_idx = next(
        (i for i, line in enumerate(lines[:10]) if _SPP_KEY_COLUMN in line), None
    )
    if header_idx is None:
        logger.debug("spp queue: header row not found")
        return []

    reader = csv.DictReader(io.StringIO("\n".join(lines[header_idx:])))
    rows: list[dict] = []
    for raw in reader:
        row = {
            key.strip(): (value or "").strip()
            for key, value in raw.items()
            if isinstance(key, str) and isinstance(value, (str, type(None)))
        }
        queue_id = row.get(_SPP_KEY_COLUMN, "")
        if not queue_id:
            continue
        rows.append(
            {
                "iso": "SPP",
                "queue_id": queue_id,
                "county": row.get("Nearest Town or County", ""),
                "state": row.get("State", ""),
                "mw": _mw(row.get("MAX Summer MW", "")),
                "fuel": row.get("Generation Type", "") or row.get("Fuel Type", ""),
                "status": row.get("Status", ""),
                "queue_date": _iso_date(row.get("Request Received", "")),
                "withdrawn": bool(row.get("Date Withdrawn", "")),
            }
        )
    return rows


def _cell_str(value: object) -> str:
    """xlsx cell -> str: None -> "", date/datetime cells -> YYYY-MM-DD,
    everything else stringified and stripped verbatim."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _sheet_grid(xlsx_bytes: bytes, sheet_name: str) -> list[tuple] | None:
    """Workbook bytes -> the named sheet's rows (values only). Anything
    that is not a readable workbook carrying that sheet -> None (degrade,
    never crash — same posture as the JSON/CSV parsers)."""
    try:
        book = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception:  # openpyxl raises BadZipFile/KeyError/... on garbage
        logger.debug("xlsx: unreadable workbook bytes")
        return None
    try:
        if sheet_name not in book.sheetnames:
            logger.debug("xlsx: sheet %r not found", sheet_name)
            return None
        return [tuple(row) for row in book[sheet_name].iter_rows(values_only=True)]
    finally:
        book.close()


def _header_columns(
    grid: list[tuple], key_column: str, scan_limit: int = 60
) -> tuple[int, dict[str, int]] | None:
    """Locate the header row by its key column (both workbooks bury it
    under a notes/banner preamble — ERCOT row 31, CAISO row 4) and return
    (header row index, {stripped header name: column index})."""
    for i, row in enumerate(grid[:scan_limit]):
        names = {c.strip(): j for j, c in enumerate(row) if isinstance(c, str) and c.strip()}
        if key_column in names:
            return i, names
    return None


def _cell(row: tuple, columns: dict[str, int], name: str) -> object:
    """Cell by header name; a missing column or a short row -> None."""
    i = columns.get(name)
    return row[i] if i is not None and i < len(row) else None


def parse_ercot(xlsx_bytes: bytes) -> list[dict]:
    """ERCOT GIS Report workbook -> normalized rows, from sheet "Project
    Details - Large Gen" (header row 31, located by its "INR" cell; the
    wrapped-header continuation rows have no INR and are skipped like any
    INR-less row). state is always "TX"; withdrawn is always False (the
    sheet lists active projects only — inactive ones live on a separate
    sheet); status is "GIM Study Phase" verbatim; queue_date carries
    "Projected COD" because the sheet has no request-received date (the
    sheet's documented 1900-01-01 placeholder dates pass through
    verbatim). ERCOT rows ALONE carry three additive keys — air_permit,
    ghg_permit, water_availability — verbatim strings, "" when blank,
    YYYY-MM-DD when the cell is a date. Unreadable bytes, a missing
    sheet, or a missing header row -> []."""
    grid = _sheet_grid(xlsx_bytes, _ERCOT_SHEET)
    if grid is None:
        return []
    found = _header_columns(grid, "INR")
    if found is None:
        logger.debug("ercot queue: header row not found")
        return []
    header_idx, columns = found
    rows: list[dict] = []
    for raw in grid[header_idx + 1 :]:
        queue_id = _cell_str(_cell(raw, columns, "INR"))
        if not queue_id:
            continue
        rows.append(
            {
                "iso": "ERCOT",
                "queue_id": queue_id,
                "county": _cell_str(_cell(raw, columns, "County")),
                "state": "TX",
                "mw": _mw(_cell(raw, columns, "Capacity (MW)")),
                "fuel": _cell_str(_cell(raw, columns, "Fuel")),
                "status": _cell_str(_cell(raw, columns, "GIM Study Phase")),
                "queue_date": _cell_str(_cell(raw, columns, "Projected COD")),
                "withdrawn": False,
                "air_permit": _cell_str(_cell(raw, columns, "Air Permit")),
                "ghg_permit": _cell_str(_cell(raw, columns, "GHG Permit")),
                "water_availability": _cell_str(_cell(raw, columns, "Water Availability")),
            }
        )
    return rows


def parse_caiso(xlsx_bytes: bytes) -> list[dict]:
    """CAISO PublicQueueReport.xlsx -> normalized rows, from sheet "Grid
    GenerationQueue" (header row 4, located by its "Queue Position" cell).
    queue_id = "Queue Position" (ints and strings like "643R" both
    appear); legend/footer rows have no Queue Position and are skipped.
    mw = "Net MWs to Grid", falling back to "MW-1" when blank; fuel =
    "Fuel-1"; queue_date = "Queue Date" (a real filing date — verified
    present in the live workbook). county/state are verbatim: counties
    come UPPERCASE and the feed spans CA/NV/AZ/MX. withdrawn mirrors
    Application Status == "WITHDRAWN" — all-False in practice, since
    withdrawn projects live on a separate sheet."""
    grid = _sheet_grid(xlsx_bytes, _CAISO_SHEET)
    if grid is None:
        return []
    found = _header_columns(grid, "Queue Position")
    if found is None:
        logger.debug("caiso queue: header row not found")
        return []
    header_idx, columns = found
    rows: list[dict] = []
    for raw in grid[header_idx + 1 :]:
        queue_id = _cell_str(_cell(raw, columns, "Queue Position"))
        if not queue_id:
            continue
        status = _cell_str(_cell(raw, columns, "Application Status"))
        mw = _cell(raw, columns, "Net MWs to Grid")
        if mw is None or (isinstance(mw, str) and not mw.strip()):
            mw = _cell(raw, columns, "MW-1")
        rows.append(
            {
                "iso": "CAISO",
                "queue_id": queue_id,
                "county": _cell_str(_cell(raw, columns, "County")),
                "state": _cell_str(_cell(raw, columns, "State")),
                "mw": _mw(mw),
                "fuel": _cell_str(_cell(raw, columns, "Fuel-1")),
                "status": status,
                "queue_date": _cell_str(_cell(raw, columns, "Queue Date")),
                "withdrawn": status.upper() == "WITHDRAWN",
            }
        )
    return rows


def join_jurisdictions(
    rows: list[dict], mapping: dict[str, str]
) -> tuple[list[dict], list[dict]]:
    """Exact-match join on f"{county}, {state}" against the explicit
    mapping table. Hits are COPIES with a "jurisdiction" key added; every
    other row is returned in misses unchanged — logged, never guessed."""
    hits: list[dict] = []
    misses: list[dict] = []
    for row in rows:
        key = f"{row.get('county', '')}, {row.get('state', '')}"
        jurisdiction = mapping.get(key)
        if jurisdiction is None:
            misses.append(row)
            continue
        hit = dict(row)
        hit["jurisdiction"] = jurisdiction
        hits.append(hit)
    if misses:
        logger.debug("jurisdiction join: %d hits, %d honest misses", len(hits), len(misses))
    return hits, misses


def load_mapping(path: str | Path = MAPPING_PATH) -> dict[str, str]:
    """Load data/iso_jurisdictions.json; missing file -> {} (keyless-
    deploy posture: an absent artifact degrades, never crashes)."""
    path = Path(path)
    if not path.exists():
        logger.debug("iso jurisdiction mapping absent at %s", path)
        return {}
    return json.loads(path.read_text())


# --------------------------------------------------------------------------
# Operational fetchers (live network; not unit-tested)
# --------------------------------------------------------------------------

_MAX_ATTEMPTS = 3
_BACKOFF_S = 2.0
_RETRY_AFTER_CAP_S = 120.0


def _get_with_retries(client: httpx.Client, url: str) -> httpx.Response | None:
    """GET with etiquette: 30s timeout, retries on transport errors and
    429/5xx, honoring Retry-After (capped) when the server sends one."""
    for attempt in range(1, _MAX_ATTEMPTS + 1):
        wait = _BACKOFF_S * attempt
        try:
            response = client.get(url, headers=HEADERS, timeout=30.0, follow_redirects=True)
        except httpx.HTTPError as exc:
            logger.debug("%s attempt %d: %s", url, attempt, type(exc).__name__)
        else:
            if response.status_code == 200:
                return response
            if response.status_code == 429 or response.status_code >= 500:
                retry_after = response.headers.get("Retry-After", "")
                if retry_after.strip().isdigit():
                    wait = min(float(retry_after), _RETRY_AFTER_CAP_S)
                logger.debug("%s attempt %d: HTTP %d", url, attempt, response.status_code)
            else:
                logger.debug("%s: HTTP %d, giving up", url, response.status_code)
                return None
        if attempt < _MAX_ATTEMPTS:
            time.sleep(wait)
    return None


def _snapshot(out_dir: Path, name: str, content: bytes, url: str) -> None:
    """Raw snapshot + provenance sidecar (fetch date + sha256, AC-1)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / name).write_bytes(content)
    provenance = {
        "url": url,
        "fetched_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "sha256": hashlib.sha256(content).hexdigest(),
        "bytes": len(content),
    }
    sidecar = out_dir / f"{name}.provenance.json"
    sidecar.write_text(json.dumps(provenance, indent=2) + "\n")


def fetch_miso(out_dir: str | Path = DEFAULT_OUT_DIR) -> list[dict]:
    """Fetch the MISO queue, snapshot the raw JSON under artifacts/iso/,
    return normalized rows ([] on any failure — degrade, never crash)."""
    with httpx.Client() as client:
        response = _get_with_retries(client, MISO_URL)
    if response is None:
        return []
    _snapshot(Path(out_dir), "miso_projects.json", response.content, MISO_URL)
    return parse_miso(response.text)


def fetch_spp(out_dir: str | Path = DEFAULT_OUT_DIR) -> list[dict]:
    """Fetch the SPP active-requests CSV, snapshot it under artifacts/iso/,
    return normalized rows ([] on any failure — degrade, never crash)."""
    with httpx.Client() as client:
        response = _get_with_retries(client, SPP_URL)
    if response is None:
        return []
    _snapshot(Path(out_dir), "spp_active.csv", response.content, SPP_URL)
    return parse_spp(response.text)


def _ercot_newest_doc_id(json_text: str) -> str:
    """ERCOT MIS listing JSON -> DocID of the newest GIS Report document.
    reportTypeId=15933 mixes Co-located Battery reports into the same
    listing (verified 2026-08-15), so entries are filtered to FriendlyName
    "GIS_Report*" before taking the max PublishDate (ISO-8601 strings with
    a fixed offset, so string comparison orders them). Malformed -> ""."""
    try:
        data = json.loads(json_text)
    except (TypeError, ValueError):
        return ""
    if not isinstance(data, dict):
        return ""
    listing = data.get("ListDocsByRptTypeRes", {})
    documents = listing.get("DocumentList", []) if isinstance(listing, dict) else []
    if not isinstance(documents, list):
        return ""
    best_id, best_published = "", ""
    for entry in documents:
        doc = entry.get("Document") if isinstance(entry, dict) else None
        if not isinstance(doc, dict):
            continue
        if not str(doc.get("FriendlyName") or "").startswith("GIS_Report"):
            continue
        doc_id = str(doc.get("DocID") or "").strip()
        published = str(doc.get("PublishDate") or "")
        if doc_id and (not best_id or published > best_published):
            best_id, best_published = doc_id, published
    return best_id


def fetch_ercot(out_dir: str | Path = DEFAULT_OUT_DIR) -> list[dict]:
    """Fetch the ERCOT GIS Report in two steps — the MIS document listing,
    then mirDownload of the newest GIS_Report DocID — snapshot the raw
    workbook under artifacts/iso/, return normalized rows ([] on any
    failure — degrade, never crash)."""
    with httpx.Client() as client:
        listing = _get_with_retries(client, ERCOT_LIST_URL)
        if listing is None:
            return []
        doc_id = _ercot_newest_doc_id(listing.text)
        if not doc_id:
            logger.debug("ercot queue: no GIS_Report entry in the MIS listing")
            return []
        url = ERCOT_DOWNLOAD_URL.format(doc_id=doc_id)
        response = _get_with_retries(client, url)
    if response is None:
        return []
    _snapshot(Path(out_dir), "ercot_gis.xlsx", response.content, url)
    return parse_ercot(response.content)


def fetch_caiso(out_dir: str | Path = DEFAULT_OUT_DIR) -> list[dict]:
    """Fetch the CAISO public queue workbook, snapshot it under
    artifacts/iso/, return normalized rows ([] on any failure — degrade,
    never crash)."""
    with httpx.Client() as client:
        response = _get_with_retries(client, CAISO_URL)
    if response is None:
        return []
    _snapshot(Path(out_dir), "caiso_queue.xlsx", response.content, CAISO_URL)
    return parse_caiso(response.content)
